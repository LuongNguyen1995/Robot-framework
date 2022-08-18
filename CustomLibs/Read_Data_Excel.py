import openpyxl
wk = openpyxl.load_workbook("C://Users//ndluong//Documents//Robot-framework//Data//Excel//Login_Register.xlsx")



def max_number_of_rows(Sheetname):
    sh = wk[Sheetname]
    return sh.max_row

def max_number_of_cols(Sheetname):
    sh = wk[Sheetname]
    return sh.max_column

def cell_data(Sheetname, row, col):
    sh = wk[Sheetname]
    cell = sh.cell(int(row),int(col))
    return cell.value


